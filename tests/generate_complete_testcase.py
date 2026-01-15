"""
Script untuk generate dokumen test case LENGKAP dalam format xlsx
Berisi semua 81 test cases yang diimplementasi di Selenium
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def create_complete_testcase_document():
    wb = Workbook()
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["No", "TC_ID", "Test Class", "Test Case Name", "Description", "Test Type", 
               "Priority", "Pre-Condition", "Test Steps", "Test Data", 
               "Expected Result", "Actual Result", "Status"]
    
    # ==================== SHEET 1: LOGIN TEST CASES (38 Tests) ====================
    ws_login = wb.active
    ws_login.title = "Login Tests (38)"
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_login.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    login_testcases = [
        # ===== TestLoginFunctionalPositive (2 tests) =====
        [1, "TC_LGN_001", "TestLoginFunctionalPositive", "test_login_valid_credentials",
         "Verify successful login with valid username and password",
         "Functional - Positive", "Critical",
         "User 'irul' registered in database with password 'irul123'",
         "1. Navigate to login.php\n2. Enter username 'irul'\n3. Enter password 'irul123'\n4. Click Sign In",
         "Username: irul\nPassword: irul123",
         "User logged in, redirected to index.php",
         "Redirected to index.php", "PASSED"],
        
        [2, "TC_LGN_002", "TestLoginFunctionalPositive", "test_login_case_sensitivity_username",
         "Verify username case sensitivity",
         "Functional - Positive", "Medium",
         "User 'irul' exists (lowercase)",
         "1. Navigate to login.php\n2. Enter 'IRUL' (uppercase)\n3. Enter correct password\n4. Click Sign In",
         "Username: IRUL\nPassword: irul123",
         "Login should fail (username case-sensitive)",
         "Login succeeded (NOT case-sensitive)", "FAILED"],
        
        # ===== TestLoginFunctionalNegative (7 tests) =====
        [3, "TC_LGN_003", "TestLoginFunctionalNegative", "test_login_invalid_password",
         "Verify login fails with wrong password",
         "Functional - Negative", "Critical",
         "User 'irul' registered in database",
         "1. Navigate to login.php\n2. Enter valid username\n3. Enter wrong password\n4. Click Sign In",
         "Username: irul\nPassword: wrongpassword",
         "Login fails, stays on login page",
         "Stayed on login.php", "PASSED"],
        
        [4, "TC_LGN_004", "TestLoginFunctionalNegative", "test_login_invalid_username",
         "Verify login fails with non-existent username",
         "Functional - Negative", "Critical",
         "Username does not exist in database",
         "1. Navigate to login.php\n2. Enter non-existent username\n3. Enter any password\n4. Click Sign In",
         "Username: nonexistentuser\nPassword: anypassword",
         "Error message: 'Register User Gagal !!'",
         "Error displayed correctly", "PASSED"],
        
        [5, "TC_LGN_005", "TestLoginFunctionalNegative", "test_login_empty_username",
         "Verify validation for empty username",
         "Functional - Negative", "High",
         "Login page accessible",
         "1. Navigate to login.php\n2. Leave username empty\n3. Enter password\n4. Click Sign In",
         "Username: (empty)\nPassword: test123",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        [6, "TC_LGN_006", "TestLoginFunctionalNegative", "test_login_empty_password",
         "Verify validation for empty password",
         "Functional - Negative", "High",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter username\n3. Leave password empty\n4. Click Sign In",
         "Username: irul\nPassword: (empty)",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        [7, "TC_LGN_007", "TestLoginFunctionalNegative", "test_login_both_fields_empty",
         "Verify validation when all fields empty",
         "Functional - Negative", "High",
         "Login page accessible",
         "1. Navigate to login.php\n2. Leave all fields empty\n3. Click Sign In",
         "Username: (empty)\nPassword: (empty)",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        [8, "TC_LGN_008", "TestLoginFunctionalNegative", "test_login_whitespace_only_username",
         "Verify validation for whitespace-only username",
         "Functional - Negative", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter spaces only in username\n3. Enter password\n4. Click Sign In",
         "Username: '   ' (spaces)\nPassword: test123",
         "Error: 'Data tidak boleh kosong !!'",
         "Handled correctly", "PASSED"],
        
        [9, "TC_LGN_009", "TestLoginFunctionalNegative", "test_login_whitespace_only_password",
         "Verify validation for whitespace-only password",
         "Functional - Negative", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter username\n3. Enter spaces only in password\n4. Click Sign In",
         "Username: irul\nPassword: '   ' (spaces)",
         "Login fails or error message",
         "Stayed on login.php", "PASSED"],
        
        # ===== TestLoginBoundaryValue (5 tests) =====
        [10, "TC_LGN_010", "TestLoginBoundaryValue", "test_login_username_min_length",
         "Verify login with 1 character username",
         "Boundary Value", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter 1-char username\n3. Enter password\n4. Click Sign In",
         "Username: a\nPassword: test123",
         "System handles 1-char username appropriately",
         "Handled without crash", "PASSED"],
        
        [11, "TC_LGN_011", "TestLoginBoundaryValue", "test_login_username_max_length",
         "Verify login with max length username (50 chars)",
         "Boundary Value", "Medium",
         "Database username field is VARCHAR(50)",
         "1. Navigate to login.php\n2. Enter 50-char username\n3. Enter password\n4. Click Sign In",
         "Username: a*50 (50 chars)\nPassword: test123",
         "System handles max-length username",
         "Handled without crash", "PASSED"],
        
        [12, "TC_LGN_012", "TestLoginBoundaryValue", "test_login_username_exceed_max",
         "Verify system handles oversized username (100 chars)",
         "Boundary Value", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter 100-char username\n3. Enter password\n4. Click Sign In",
         "Username: a*100 (100 chars)\nPassword: test123",
         "System handles gracefully without crash",
         "No crash occurred", "PASSED"],
        
        [13, "TC_LGN_013", "TestLoginBoundaryValue", "test_login_password_min_length",
         "Verify login with 1 character password",
         "Boundary Value", "Low",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter username\n3. Enter 1-char password\n4. Click Sign In",
         "Username: irul\nPassword: a",
         "Login fails (incorrect password)",
         "Login failed as expected", "PASSED"],
        
        [14, "TC_LGN_014", "TestLoginBoundaryValue", "test_login_password_very_long",
         "Verify system handles very long password (300 chars)",
         "Boundary Value", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter username\n3. Enter 300-char password\n4. Click Sign In",
         "Username: irul\nPassword: a*300 (300 chars)",
         "System handles without crash",
         "No crash occurred", "PASSED"],
        
        # ===== TestLoginSecurity (7 tests) =====
        [15, "TC_LGN_015", "TestLoginSecurity", "test_sql_injection_basic",
         "Verify protection against basic SQL injection",
         "Security", "Critical",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter SQL injection\n3. Enter any password\n4. Click Sign In",
         "Username: ' OR '1'='1\nPassword: test",
         "Login fails, SQL injection blocked",
         "LOGIN SUCCEEDED - VULNERABLE!", "FAILED"],
        
        [16, "TC_LGN_016", "TestLoginSecurity", "test_sql_injection_comment",
         "Verify protection against SQL comment injection",
         "Security", "Critical",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter SQL with comment\n3. Enter password\n4. Click Sign In",
         "Username: admin'--\nPassword: anything",
         "Login fails, SQL injection blocked",
         "Login failed", "PASSED"],
        
        [17, "TC_LGN_017", "TestLoginSecurity", "test_sql_injection_union",
         "Verify protection against UNION-based injection",
         "Security", "Critical",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter UNION injection\n3. Enter password\n4. Click Sign In",
         "Username: ' UNION SELECT * FROM users--\nPassword: test",
         "Login fails, no data leakage",
         "Login failed", "PASSED"],
        
        [18, "TC_LGN_018", "TestLoginSecurity", "test_sql_injection_password",
         "Verify SQL injection protection in password field",
         "Security", "Critical",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter valid username\n3. Enter SQL injection in password\n4. Click Sign In",
         "Username: irul\nPassword: ' OR '1'='1",
         "Login fails (password is hashed)",
         "Login failed", "PASSED"],
        
        [19, "TC_LGN_019", "TestLoginSecurity", "test_xss_stored_username",
         "Verify XSS protection in username field",
         "Security", "Critical",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter XSS script in username\n3. Click Sign In",
         "Username: <script>alert('XSS')</script>\nPassword: test",
         "Script not executed, input sanitized",
         "No alert triggered", "PASSED"],
        
        [20, "TC_LGN_020", "TestLoginSecurity", "test_xss_event_handler",
         "Verify XSS protection against event handlers",
         "Security", "High",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter XSS with event handler\n3. Click Sign In",
         "Username: <img src=x onerror=alert('XSS')>\nPassword: test",
         "Script not executed",
         "No alert triggered", "PASSED"],
        
        [21, "TC_LGN_021", "TestLoginSecurity", "test_html_injection",
         "Verify HTML injection protection",
         "Security", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter HTML tags in username\n3. Click Sign In",
         "Username: <h1>Hacked</h1>\nPassword: test",
         "HTML not rendered, input escaped",
         "HTML escaped/not rendered", "PASSED"],
        
        # ===== TestLoginSpecialCharacters (3 tests) =====
        [22, "TC_LGN_022", "TestLoginSpecialCharacters", "test_username_special_chars",
         "Verify handling of special characters in username",
         "Special Characters", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter username with special chars\n3. Click Sign In",
         "Username: user@#$%\nPassword: test123",
         "System handles special characters appropriately",
         "Handled without crash", "PASSED"],
        
        [23, "TC_LGN_023", "TestLoginSpecialCharacters", "test_username_unicode",
         "Verify handling of unicode characters",
         "Special Characters", "Low",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter unicode characters\n3. Click Sign In",
         "Username: 用户名\nPassword: test123",
         "System handles unicode appropriately",
         "Handled without crash", "PASSED"],
        
        [24, "TC_LGN_024", "TestLoginSpecialCharacters", "test_password_special_chars",
         "Verify login with special chars in password",
         "Special Characters", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Enter username\n3. Enter password with special chars\n4. Click Sign In",
         "Username: irul\nPassword: P@ss!@#$%^&*()",
         "Login fails (wrong password) but no crash",
         "Failed gracefully", "PASSED"],
        
        # ===== TestLoginUsability (4 tests) =====
        [25, "TC_LGN_025", "TestLoginUsability", "test_tab_navigation",
         "Verify form can be navigated using Tab key",
         "Usability", "Low",
         "Login page accessible",
         "1. Navigate to login.php\n2. Click username field\n3. Press Tab",
         "N/A",
         "Focus moves to password field",
         "Tab navigation works", "PASSED"],
        
        [26, "TC_LGN_026", "TestLoginUsability", "test_enter_key_submission",
         "Verify form submits on Enter key",
         "Usability", "Medium",
         "Login page accessible",
         "1. Fill in credentials\n2. Press Enter instead of clicking button",
         "Username: irul\nPassword: irul123",
         "Form submits and login processed",
         "Form submitted on Enter", "PASSED"],
        
        [27, "TC_LGN_027", "TestLoginUsability", "test_password_masking",
         "Verify password field masks input",
         "Usability", "Medium",
         "Login page accessible",
         "1. Navigate to login.php\n2. Check password field type",
         "N/A",
         "Password field type is 'password'",
         "Type is 'password'", "PASSED"],
        
        [28, "TC_LGN_028", "TestLoginUsability", "test_page_elements_present",
         "Verify all expected elements present",
         "Usability", "Low",
         "Login page accessible",
         "1. Navigate to login.php\n2. Check for username, password, submit, register link",
         "N/A",
         "All elements present",
         "All elements found", "PASSED"],
        
        # ===== TestLoginDatabaseEdgeCases (2 tests) =====
        [29, "TC_LGN_029", "TestLoginDatabaseEdgeCases", "test_login_user_with_empty_name_in_db",
         "Verify login works when name field is empty in DB",
         "Database - Edge Case", "High",
         "User 'irul' has empty name field in database",
         "1. Navigate to login.php\n2. Login with user having empty name\n3. Click Sign In",
         "Username: irul\nPassword: irul123",
         "Login successful despite empty name field",
         "Login succeeded", "PASSED"],
        
        [30, "TC_LGN_030", "TestLoginDatabaseEdgeCases", "test_login_user_ahmad_empty_name",
         "Verify another user with empty name can login",
         "Database - Edge Case", "Medium",
         "User 'ahmad' has empty name field in database",
         "1. Navigate to login.php\n2. Login with ahmad\n3. Click Sign In",
         "Username: ahmad\nPassword: ahmad123",
         "Login successful despite empty name field",
         "Login succeeded", "PASSED"],
        
        # ===== TestLoginSession (1 test) =====
        [31, "TC_LGN_031", "TestLoginSession", "test_already_logged_in_redirect",
         "Verify redirect when already logged in",
         "Session", "Medium",
         "User has active session",
         "1. Login successfully\n2. Navigate to login.php again",
         "N/A",
         "Should redirect to index.php",
         "Redirected correctly", "PASSED"],
        
        # ===== TestLoginWithStubProfessional (7 tests) =====
        [32, "TC_LGN_032", "TestLoginWithStubProfessional", "test_stub_valid_login",
         "Stub: Valid login credentials",
         "Unit Test - Stub", "Critical",
         "DatabaseStub initialized",
         "1. Call attempt_login with valid credentials",
         "Username: irul\nPassword: irul123",
         "success=True, redirect='index.php'",
         "Stub returned success", "PASSED"],
        
        [33, "TC_LGN_033", "TestLoginWithStubProfessional", "test_stub_invalid_username",
         "Stub: Non-existent username",
         "Unit Test - Stub", "High",
         "DatabaseStub initialized",
         "1. Call attempt_login with non-existent user",
         "Username: nonexistent\nPassword: password",
         "success=False, error contains 'Gagal'",
         "Stub returned failure", "PASSED"],
        
        [34, "TC_LGN_034", "TestLoginWithStubProfessional", "test_stub_empty_username",
         "Stub: Empty username",
         "Unit Test - Stub", "High",
         "DatabaseStub initialized",
         "1. Call attempt_login with empty username",
         "Username: (empty)\nPassword: password",
         "success=False, error contains 'kosong'",
         "Stub returned failure", "PASSED"],
        
        [35, "TC_LGN_035", "TestLoginWithStubProfessional", "test_stub_empty_password",
         "Stub: Empty password",
         "Unit Test - Stub", "High",
         "DatabaseStub initialized",
         "1. Call attempt_login with empty password",
         "Username: irul\nPassword: (empty)",
         "success=False, error contains 'kosong'",
         "Stub returned failure", "PASSED"],
        
        [36, "TC_LGN_036", "TestLoginWithStubProfessional", "test_stub_whitespace_username",
         "Stub: Whitespace-only username",
         "Unit Test - Stub", "Medium",
         "DatabaseStub initialized",
         "1. Call attempt_login with whitespace username",
         "Username: '   '\nPassword: password",
         "success=False",
         "Stub returned failure", "PASSED"],
        
        [37, "TC_LGN_037", "TestLoginWithStubProfessional", "test_stub_user_with_empty_name",
         "Stub: Verify user with empty name can login",
         "Unit Test - Stub", "High",
         "DatabaseStub with user having empty name",
         "1. Get user from stub\n2. Verify name is empty\n3. Attempt login",
         "Username: irul",
         "User name is empty, login succeeds",
         "Name empty, login success", "PASSED"],
        
        [38, "TC_LGN_038", "TestLoginWithStubProfessional", "test_stub_sql_injection_attempt",
         "Stub: SQL injection should fail",
         "Unit Test - Stub", "Critical",
         "DatabaseStub initialized",
         "1. Call attempt_login with SQL injection",
         "Username: ' OR '1'='1\nPassword: test",
         "success=False",
         "Stub returned failure", "PASSED"],
    ]
    
    # Write login test cases
    for row_idx, tc in enumerate(login_testcases, 2):
        for col_idx, value in enumerate(tc, 1):
            cell = ws_login.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
            # Color status column
            if col_idx == 13:  # Status column
                if value == "PASSED":
                    cell.fill = pass_fill
                elif value == "FAILED":
                    cell.fill = fail_fill
    
    # Set column widths
    login_col_widths = [4, 12, 28, 35, 40, 18, 10, 30, 40, 30, 35, 30, 10]
    for i, width in enumerate(login_col_widths, 1):
        ws_login.column_dimensions[get_column_letter(i)].width = width
    
    # ==================== SHEET 2: REGISTER TEST CASES (43 Tests) ====================
    ws_register = wb.create_sheet("Register Tests (43)")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_register.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    register_testcases = [
        # ===== TestRegisterFunctionalPositive (2 tests) =====
        [1, "TC_REG_001", "TestRegisterFunctionalPositive", "test_register_valid_data",
         "Verify successful registration with all valid data",
         "Functional - Positive", "Critical",
         "Database accessible, username not taken",
         "1. Navigate to register.php\n2. Fill all fields\n3. Click Register",
         "Name: Test User\nEmail: test@test.com\nUsername: unique\nPassword: Test@123\nRe-Password: Test@123",
         "Registration successful, redirect to index.php",
         "Registration processed", "PASSED"],
        
        [2, "TC_REG_002", "TestRegisterFunctionalPositive", "test_register_minimum_valid_data",
         "Verify registration with minimum required data",
         "Functional - Positive", "High",
         "Database accessible",
         "1. Fill required fields with minimum valid data\n2. Submit",
         "Name: A\nEmail: a@b.co\nUsername: ab\nPassword: 1\nRe-Password: 1",
         "Registration processed",
         "Processed without crash", "PASSED"],
        
        # ===== TestRegisterEmptyFields (6 tests) =====
        [3, "TC_REG_003", "TestRegisterEmptyFields", "test_register_empty_name",
         "Verify validation for empty name",
         "Functional - Negative", "High",
         "Register page accessible",
         "1. Leave name empty\n2. Fill other fields\n3. Click Register",
         "Name: (empty)\nOther fields: valid",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        [4, "TC_REG_004", "TestRegisterEmptyFields", "test_register_empty_email",
         "Verify validation for empty email",
         "Functional - Negative", "High",
         "Register page accessible",
         "1. Fill name\n2. Leave email empty\n3. Fill other fields\n4. Click Register",
         "Email: (empty)\nOther fields: valid",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        [5, "TC_REG_005", "TestRegisterEmptyFields", "test_register_empty_username",
         "Verify validation for empty username",
         "Functional - Negative", "High",
         "Register page accessible",
         "1. Fill name, email\n2. Leave username empty\n3. Fill password\n4. Click Register",
         "Username: (empty)\nOther fields: valid",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        [6, "TC_REG_006", "TestRegisterEmptyFields", "test_register_empty_password",
         "Verify validation for empty password",
         "Functional - Negative", "High",
         "Register page accessible",
         "1. Fill name, email, username\n2. Leave password empty\n3. Click Register",
         "Password: (empty)\nOther fields: valid",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        [7, "TC_REG_007", "TestRegisterEmptyFields", "test_register_empty_repassword",
         "Verify validation for empty re-password",
         "Functional - Negative", "High",
         "Register page accessible",
         "1. Fill all except re-password\n2. Click Register",
         "Re-Password: (empty)\nOther fields: valid",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        [8, "TC_REG_008", "TestRegisterEmptyFields", "test_register_all_fields_empty",
         "Verify validation when all fields empty",
         "Functional - Negative", "High",
         "Register page accessible",
         "1. Leave all fields empty\n2. Click Register",
         "All fields: (empty)",
         "Error: 'Data tidak boleh kosong !!'",
         "Error displayed", "PASSED"],
        
        # ===== TestRegisterPasswordValidation (2 tests) =====
        [9, "TC_REG_009", "TestRegisterPasswordValidation", "test_register_password_mismatch",
         "Verify validation when passwords don't match",
         "Validation", "Critical",
         "Register page accessible",
         "1. Fill all fields\n2. Enter different passwords\n3. Click Register",
         "Password: password1\nRe-Password: password2",
         "Error: 'Password tidak sama !!'",
         "Error displayed", "PASSED"],
        
        [10, "TC_REG_010", "TestRegisterPasswordValidation", "test_register_password_case_difference",
         "Verify case sensitivity in password match",
         "Validation", "Medium",
         "Register page accessible",
         "1. Enter password with different case in re-password",
         "Password: Password123\nRe-Password: password123",
         "Error: 'Password tidak sama !!'",
         "Error displayed", "PASSED"],
        
        # ===== TestRegisterDuplicateValidation (1 test) =====
        [11, "TC_REG_011", "TestRegisterDuplicateValidation", "test_register_duplicate_name_field",
         "Verify rejection when name matches existing username (BUG)",
         "Validation - Duplicate", "Critical",
         "Username 'irul' exists in database",
         "1. Enter 'irul' in name field\n2. Enter unique username\n3. Click Register",
         "Name: irul\nUsername: newuniqueuser",
         "BUG: Shows 'Username sudah terdaftar' even though username is unique",
         "Bug confirmed", "PASSED"],
        
        # ===== TestRegisterEmailValidation (3 tests) =====
        [12, "TC_REG_012", "TestRegisterEmailValidation", "test_register_invalid_email_no_at",
         "Verify email format validation (no @)",
         "Validation - Email", "High",
         "Register page accessible",
         "1. Enter email without @ symbol\n2. Submit",
         "Email: invalidemail.com",
         "Form validation error (HTML5)",
         "Stayed on register page", "PASSED"],
        
        [13, "TC_REG_013", "TestRegisterEmailValidation", "test_register_invalid_email_no_domain",
         "Verify email format validation (no domain)",
         "Validation - Email", "Medium",
         "Register page accessible",
         "1. Enter email without domain\n2. Submit",
         "Email: test@",
         "Form validation error",
         "Stayed on register page", "PASSED"],
        
        [14, "TC_REG_014", "TestRegisterEmailValidation", "test_register_valid_email_subdomain",
         "Verify email with subdomain accepted",
         "Validation - Email", "Low",
         "Register page accessible",
         "1. Enter email with subdomain\n2. Submit",
         "Email: test@mail.domain.com",
         "Email accepted as valid",
         "Processed successfully", "PASSED"],
        
        # ===== TestRegisterBoundaryValue (6 tests) =====
        [15, "TC_REG_015", "TestRegisterBoundaryValue", "test_register_name_min_length",
         "Verify minimum name length (1 char)",
         "Boundary Value", "Medium",
         "Register page accessible",
         "1. Enter 1-character name\n2. Fill other fields\n3. Submit",
         "Name: A",
         "Accepted or rejected based on requirement",
         "Handled without crash", "PASSED"],
        
        [16, "TC_REG_016", "TestRegisterBoundaryValue", "test_register_name_max_length",
         "Verify maximum name length (70 chars)",
         "Boundary Value", "Medium",
         "Database name field is VARCHAR(70)",
         "1. Enter 70-character name\n2. Submit",
         "Name: A*70 (70 chars)",
         "Name accepted",
         "Handled without crash", "PASSED"],
        
        [17, "TC_REG_017", "TestRegisterBoundaryValue", "test_register_name_exceed_max",
         "Verify handling of oversized name (100 chars)",
         "Boundary Value", "Medium",
         "Register page accessible",
         "1. Enter 100-character name\n2. Submit",
         "Name: A*100 (100 chars)",
         "Truncated or error message",
         "No crash occurred", "PASSED"],
        
        [18, "TC_REG_018", "TestRegisterBoundaryValue", "test_register_username_min_length",
         "Verify minimum username length (1 char)",
         "Boundary Value", "Medium",
         "Register page accessible",
         "1. Enter 1-character username\n2. Submit",
         "Username: a",
         "Accepted or rejected based on requirement",
         "Handled without crash", "PASSED"],
        
        [19, "TC_REG_019", "TestRegisterBoundaryValue", "test_register_username_exceed_max",
         "Verify handling of oversized username (100 chars)",
         "Boundary Value", "Medium",
         "Register page accessible",
         "1. Enter 100-character username\n2. Submit",
         "Username: a*100 (100 chars)",
         "Truncated or error message",
         "No crash occurred", "PASSED"],
        
        [20, "TC_REG_020", "TestRegisterBoundaryValue", "test_register_password_very_long",
         "Verify handling of very long password (300 chars)",
         "Boundary Value", "Medium",
         "Register page accessible",
         "1. Enter 300-character password\n2. Submit",
         "Password: a*300 (300 chars)\nRe-Password: same",
         "System handles without crash",
         "No crash occurred", "PASSED"],
        
        # ===== TestRegisterSecurity (6 tests) =====
        [21, "TC_REG_021", "TestRegisterSecurity", "test_sql_injection_name_field",
         "Verify SQL injection protection in name",
         "Security", "Critical",
         "Register page accessible",
         "1. Enter SQL injection in name field\n2. Submit",
         "Name: ' OR '1'='1",
         "Registration fails or input sanitized",
         "Handled safely", "PASSED"],
        
        [22, "TC_REG_022", "TestRegisterSecurity", "test_sql_injection_username_field",
         "Verify SQL injection protection in username",
         "Security", "Critical",
         "Register page accessible",
         "1. Enter SQL injection in username\n2. Submit",
         "Username: '; DROP TABLE users;--",
         "SQL not executed, table intact",
         "Table intact", "PASSED"],
        
        [23, "TC_REG_023", "TestRegisterSecurity", "test_sql_injection_email_field",
         "Verify SQL injection protection in email",
         "Security", "Critical",
         "Register page accessible",
         "1. Enter SQL injection in email\n2. Submit",
         "Email: test@test.com'; DELETE FROM users;--",
         "SQL not executed, input sanitized",
         "Handled safely", "PASSED"],
        
        [24, "TC_REG_024", "TestRegisterSecurity", "test_xss_attack_name_field",
         "Verify XSS protection in name",
         "Security", "Critical",
         "Register page accessible",
         "1. Enter XSS script in name\n2. Submit",
         "Name: <script>alert('XSS')</script>",
         "Script not executed, input sanitized",
         "No alert triggered", "PASSED"],
        
        [25, "TC_REG_025", "TestRegisterSecurity", "test_xss_attack_username_field",
         "Verify XSS protection in username",
         "Security", "High",
         "Register page accessible",
         "1. Enter XSS in username\n2. Submit",
         "Username: <script>document.location='evil.com'</script>",
         "Script not executed",
         "No alert triggered", "PASSED"],
        
        [26, "TC_REG_026", "TestRegisterSecurity", "test_html_injection",
         "Verify HTML injection protection",
         "Security", "Medium",
         "Register page accessible",
         "1. Enter HTML tags in form fields\n2. Submit",
         "Name: <b>Bold</b><h1>Header</h1>",
         "HTML tags escaped, not rendered",
         "HTML not rendered", "PASSED"],
        
        # ===== TestRegisterSpecialCharacters (3 tests) =====
        [27, "TC_REG_027", "TestRegisterSpecialCharacters", "test_name_with_special_chars",
         "Verify handling special chars in name",
         "Special Characters", "Medium",
         "Register page accessible",
         "1. Enter name with special characters\n2. Submit",
         "Name: John O'Brien-Smith Jr.",
         "Name accepted",
         "Handled without crash", "PASSED"],
        
        [28, "TC_REG_028", "TestRegisterSpecialCharacters", "test_username_underscore_dash",
         "Verify common username characters",
         "Special Characters", "Medium",
         "Register page accessible",
         "1. Enter username with underscore/dash\n2. Submit",
         "Username: john_doe-123",
         "Username accepted",
         "Handled without crash", "PASSED"],
        
        [29, "TC_REG_029", "TestRegisterSpecialCharacters", "test_password_all_special_chars",
         "Verify strong password accepted",
         "Special Characters", "Medium",
         "Register page accessible",
         "1. Enter password with special characters\n2. Submit",
         "Password: P@$$w0rd!@#$%^&*()\nRe-Password: same",
         "Password accepted and properly hashed",
         "Handled without crash", "PASSED"],
        
        # ===== TestRegisterUsability (5 tests) =====
        [30, "TC_REG_030", "TestRegisterUsability", "test_navigation_to_login",
         "Verify link to login page works",
         "Usability - Navigation", "Medium",
         "Register page accessible",
         "1. Click 'Login' link on register page",
         "N/A",
         "Redirect to login.php",
         "Redirected to login.php", "PASSED"],
        
        [31, "TC_REG_031", "TestRegisterUsability", "test_tab_order_navigation",
         "Verify logical tab order",
         "Usability", "Low",
         "Register page accessible",
         "1. Press Tab key through all fields",
         "N/A",
         "Tab order: Name → Email → Username → Password → Re-Password",
         "Tab order correct", "PASSED"],
        
        [32, "TC_REG_032", "TestRegisterUsability", "test_form_labels_present",
         "Verify all fields have labels",
         "Usability - Accessibility", "Low",
         "Register page accessible",
         "1. Check each field has associated label",
         "N/A",
         "All fields have descriptive labels",
         "Labels present", "PASSED"],
        
        [33, "TC_REG_033", "TestRegisterUsability", "test_placeholder_text",
         "Verify placeholder text is helpful",
         "Usability", "Low",
         "Register page accessible",
         "1. Check placeholder text in each field",
         "N/A",
         "Placeholders provide helpful hints",
         "Placeholders present", "PASSED"],
        
        [34, "TC_REG_034", "TestRegisterUsability", "test_error_message_visibility",
         "Verify error messages are visible",
         "Usability", "Medium",
         "Register page accessible",
         "1. Trigger error\n2. Check error message display",
         "Submit with empty fields",
         "Error message clearly visible",
         "Error visible", "PASSED"],
        
        # ===== TestRegisterBugVerification (2 tests) =====
        [35, "TC_REG_035", "TestRegisterBugVerification", "test_bug_variable_name_mismatch",
         "Verify $nama vs $name bug - name won't be saved",
         "Bug Verification", "Critical",
         "Register page accessible, database accessible",
         "1. Register with name 'Test Name'\n2. Check database after registration",
         "Name: This Name Should Be Saved",
         "BUG: Name field will be empty in database",
         "Bug confirmed (name empty in DB)", "PASSED"],
        
        [36, "TC_REG_036", "TestRegisterBugVerification", "test_bug_wrong_duplicate_check",
         "Verify cek_nama checks wrong field",
         "Bug Verification", "High",
         "User 'irul' exists in database",
         "1. Enter 'irul' in name field\n2. Enter unique username\n3. Submit",
         "Name: irul\nUsername: completelyuniqueuser123",
         "BUG: Will show 'Username sudah terdaftar' for unique username",
         "Bug confirmed", "PASSED"],
        
        # ===== TestRegisterWithStubProfessional (7 tests) =====
        [37, "TC_REG_037", "TestRegisterWithStubProfessional", "test_stub_valid_registration",
         "Stub: Valid registration data",
         "Unit Test - Stub", "Critical",
         "DatabaseStub initialized",
         "1. Call attempt_register with valid data",
         "Name: New User\nEmail: new@test.com\nUsername: newuser\nPassword: password123",
         "success=True",
         "Stub returned success", "PASSED"],
        
        [38, "TC_REG_038", "TestRegisterWithStubProfessional", "test_stub_empty_name",
         "Stub: Empty name field",
         "Unit Test - Stub", "High",
         "DatabaseStub initialized",
         "1. Call attempt_register with empty name",
         "Name: (empty)",
         "success=False, error contains 'kosong'",
         "Stub returned failure", "PASSED"],
        
        [39, "TC_REG_039", "TestRegisterWithStubProfessional", "test_stub_empty_email",
         "Stub: Empty email field",
         "Unit Test - Stub", "High",
         "DatabaseStub initialized",
         "1. Call attempt_register with empty email",
         "Email: (empty)",
         "success=False",
         "Stub returned failure", "PASSED"],
        
        [40, "TC_REG_040", "TestRegisterWithStubProfessional", "test_stub_empty_username",
         "Stub: Empty username field",
         "Unit Test - Stub", "High",
         "DatabaseStub initialized",
         "1. Call attempt_register with empty username",
         "Username: (empty)",
         "success=False",
         "Stub returned failure", "PASSED"],
        
        [41, "TC_REG_041", "TestRegisterWithStubProfessional", "test_stub_password_mismatch",
         "Stub: Password mismatch",
         "Unit Test - Stub", "Critical",
         "DatabaseStub initialized",
         "1. Call attempt_register with mismatched passwords",
         "Password: password1\nRe-Password: password2",
         "success=False, validate contains 'sama'",
         "Stub returned failure", "PASSED"],
        
        [42, "TC_REG_042", "TestRegisterWithStubProfessional", "test_stub_duplicate_name_check_bug",
         "Stub: Verify bug - checks name instead of username",
         "Unit Test - Stub", "High",
         "DatabaseStub with user 'irul'",
         "1. Call attempt_register with name='irul', unique username",
         "Name: irul\nUsername: newuniqueuser",
         "success=False, error contains 'terdaftar'",
         "Stub confirmed bug", "PASSED"],
        
        [43, "TC_REG_043", "TestRegisterWithStubProfessional", "test_stub_database_empty_name_field",
         "Stub: Verify users have empty name field in DB",
         "Unit Test - Stub", "High",
         "DatabaseStub initialized",
         "1. Get user 'irul' from stub\n2. Get user 'ahmad' from stub\n3. Check name fields",
         "N/A",
         "Both users have empty name field",
         "Name fields are empty", "PASSED"],
    ]
    
    # Write register test cases
    for row_idx, tc in enumerate(register_testcases, 2):
        for col_idx, value in enumerate(tc, 1):
            cell = ws_register.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
            # Color status column
            if col_idx == 13:  # Status column
                if value == "PASSED":
                    cell.fill = pass_fill
                elif value == "FAILED":
                    cell.fill = fail_fill
    
    # Set column widths
    for i, width in enumerate(login_col_widths, 1):
        ws_register.column_dimensions[get_column_letter(i)].width = width
    
    # ==================== SHEET 3: SUMMARY ====================
    ws_summary = wb.create_sheet("Summary")
    
    summary_data = [
        ["TEST EXECUTION SUMMARY", "", "", "", ""],
        ["", "", "", "", ""],
        ["Module", "Total Tests", "Passed", "Failed", "Pass Rate"],
        ["Login Module", "38", "36", "2", "94.7%"],
        ["Register Module", "43", "43", "0", "100%"],
        ["TOTAL", "81", "79", "2", "97.5%"],
        ["", "", "", "", ""],
        ["FAILED TESTS:", "", "", "", ""],
        ["TC_LGN_002", "test_login_case_sensitivity_username", "Username NOT case-sensitive", "Finding", ""],
        ["TC_LGN_015", "test_sql_injection_basic", "SQL Injection VULNERABLE", "Critical Bug", ""],
        ["", "", "", "", ""],
        ["TEST CATEGORIES:", "", "", "", ""],
        ["Category", "Login", "Register", "Total", ""],
        ["Functional - Positive", "2", "2", "4", ""],
        ["Functional - Negative", "7", "6", "13", ""],
        ["Boundary Value", "5", "6", "11", ""],
        ["Security", "7", "6", "13", ""],
        ["Special Characters", "3", "3", "6", ""],
        ["Usability", "4", "5", "9", ""],
        ["Validation", "0", "5", "5", ""],
        ["Database Edge Cases", "2", "0", "2", ""],
        ["Session", "1", "0", "1", ""],
        ["Bug Verification", "0", "2", "2", ""],
        ["Unit Test - Stub", "7", "7", "14", ""],
        ["", "", "", "", ""],
        ["KNOWN BUGS TESTED:", "", "", "", ""],
        ["1. register.php line 34: $nama vs $name variable mismatch (name not saved)", "", "", "", ""],
        ["2. cek_nama() function checks 'name' field instead of 'username' for duplicates", "", "", "", ""],
        ["3. Users in database have empty 'name' field (intentional for testing)", "", "", "", ""],
        ["4. SQL Injection vulnerability in login (username field)", "", "", "", ""],
        ["5. Username is NOT case-sensitive", "", "", "", ""],
        ["", "", "", "", ""],
        ["TEST EXECUTION INFO:", "", "", "", ""],
        ["Date", "January 15, 2026", "", "", ""],
        ["Platform", "Windows 11", "", "", ""],
        ["Browser", "Chrome (via Selenium WebDriver)", "", "", ""],
        ["Python Version", "3.13.0", "", "", ""],
        ["Pytest Version", "9.0.2", "", "", ""],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
            elif row_idx in [3, 6, 8, 12]:
                cell.font = Font(bold=True)
                if row_idx == 3 or row_idx == 12:
                    cell.fill = header_fill
                    cell.font = header_font
            elif row_idx == 6:  # Total row
                cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True)
            cell.border = thin_border
    
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 40
    ws_summary.column_dimensions['C'].width = 35
    ws_summary.column_dimensions['D'].width = 15
    ws_summary.column_dimensions['E'].width = 15
    
    # Save
    wb.save('TestCase_Complete_81_Tests.xlsx')
    print("Complete Test Case document created: TestCase_Complete_81_Tests.xlsx")
    print("Total Test Cases: 81 (Login: 38, Register: 43)")
    print("Passed: 79, Failed: 2")

if __name__ == "__main__":
    create_complete_testcase_document()
